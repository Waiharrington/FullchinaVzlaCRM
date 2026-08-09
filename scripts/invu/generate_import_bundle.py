#!/usr/bin/env python3
"""Build a private, import-ready bundle from a FullChina Invu export.

The output contains client data and must stay outside Git. The repository only
stores this transformer and the schema/loader scripts.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


CATEGORY_MAP = {
    "arroz": ("arroz", "🍚"),
    "raciones": ("racion", "🥡"),
    "pasta y chopsuey": ("wok", "🍜"),
    "bebidas": ("bebida", "🥤"),
    "extras": ("extra", "➕"),
    "promociones": ("plato", "🍽️"),
    "otros": ("plato", "🍽️"),
    "menu especial": ("plato", "🍽️"),
}

UNIT_SYMBOLS = {
    "kilogramo": "kg",
    "litros": "L",
    "litro": "L",
    "unidad": "und",
    "porción": "por",
    "porcion": "por",
    "gramos": "g",
    "gramo": "g",
    "mililitros": "ml",
    "mililitro": "ml",
}


def compact(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split())


def scalar(value: Any) -> Any:
    return value.get("value", "") if isinstance(value, dict) else value


def label(value: Any) -> str:
    return clean(value.get("label", "")) if isinstance(value, dict) else clean(value)


def yes(value: Any) -> bool:
    return clean(value).casefold() in {"1", "yes", "si", "sí", "true", "active", "activo"}


def number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = clean(value).replace("$", "").replace("Bs.", "").replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def integer(value: Any) -> int | None:
    try:
        return int(float(str(value)))
    except (TypeError, ValueError):
        return None


def iso(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    text = clean(value)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).isoformat()
        except ValueError:
            pass
    return ""


def plain_html(value: Any) -> str:
    text = re.sub(r"<[^>]+>", " ", str(value or ""))
    return clean(html.unescape(text))


def category_for(raw: str) -> tuple[str, str]:
    normalized = raw.casefold()
    for needle, mapped in CATEGORY_MAP.items():
        if needle in normalized:
            return mapped
    return "plato", "🍽️"


def display_name(value: Any) -> str:
    text = clean(value)
    return text.title() if text and text == text.upper() else text


def workbook_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    sheet = load_workbook(path, read_only=True, data_only=True).active
    iterator = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(iterator)]
    rows = []
    for values in iterator:
        if not any(value not in (None, "") for value in values):
            continue
        rows.append(dict(zip(headers, values)))
    return headers, rows


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fields = list(rows[0]) if rows else []
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def build(source: Path, output: Path) -> dict[str, Any]:
    output.mkdir(parents=True, exist_ok=True)
    official = source / "raw" / "official_exports"
    menu_records = json.loads((source / "structured" / "menu_items.json").read_text(encoding="utf-8"))
    ingredient_records = json.loads((source / "structured" / "inventory_products.json").read_text(encoding="utf-8"))

    _, modifier_source = workbook_rows(official / "modifiers.xlsx")
    _, option_source = workbook_rows(official / "modifier_options.xlsx")
    options_by_modifier: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in option_source:
        options_by_modifier[clean(row["Modifier code"])].append(row)

    modifier_by_item: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in modifier_source:
        assigned = clean(row["Item code to which you want to Assign (in case of assigning to several items, you can add them separating them by comma (,). EX: A1,A2,A3)"])
        for code in re.split(r"\s*,\s*", assigned):
            if code:
                modifier_by_item[code].append(row)

    products: list[dict[str, Any]] = []
    product_code_to_key: dict[str, str] = {}
    for record in menu_records:
        fields = record["fields"]
        source_id = str(record["source_id"])
        code = clean(scalar(fields.get("codigo")))
        name = display_name(scalar(fields.get("nombre")))
        price = number(scalar(fields.get("precioSugerido")))
        source_active = clean(scalar(fields.get("status", "1"))) == "1"
        category, emoji = category_for(label(fields.get("idCategoriaMenu_FK")))
        source_key = f"menu:{source_id}"
        product_code_to_key[code] = source_key
        products.append({
            "source_key": source_key,
            "source_code": code,
            "name": name,
            "description": clean(scalar(fields.get("descripcion"))),
            "price": f"{price:.2f}",
            "cost": "",
            "category": category,
            "emoji": emoji,
            "is_active": str(source_active and price > 0).lower(),
            "barcode": clean(scalar(fields.get("codigoBarra"))),
            "display_order": integer(scalar(fields.get("orden"))) or 0,
            "source_payload": compact(record),
        })

        if source_active and price <= 0:
            for modifier in modifier_by_item.get(code, []):
                modifier_code = clean(modifier["Code (If it is new, you can leave this field empty, fill if you want to modify)"])
                for option in options_by_modifier.get(modifier_code, []):
                    option_price = number(option["Sale price"])
                    if option_price <= 0 or not yes(option["Active? (Yes/No)"]):
                        continue
                    option_code = clean(option["Modifier option code"])
                    option_name = display_name(option["Modifier option name"])
                    products.append({
                        "source_key": f"{source_key}:option:{option_code}",
                        "source_code": f"{code}:{option_code}",
                        "name": f"{name} — {option_name}",
                        "description": clean(scalar(fields.get("descripcion"))),
                        "price": f"{option_price:.2f}",
                        "cost": "",
                        "category": category,
                        "emoji": emoji,
                        "is_active": "true",
                        "barcode": "",
                        "display_order": integer(scalar(fields.get("orden"))) or 0,
                        "source_payload": compact({"derived_from": record, "modifier": modifier, "option": option}),
                    })

    units_by_name: dict[str, dict[str, str]] = {}
    ingredients: list[dict[str, Any]] = []
    ingredient_code_to_key: dict[str, str] = {}
    for record in ingredient_records:
        fields = record["fields"]
        source_id = str(record["source_id"])
        code = clean(scalar(fields.get("codigo")))
        unit_field = fields.get("unidad_inventario") or fields.get("unidad")
        unit_name = label(unit_field) or "Unidad"
        units_by_name.setdefault(unit_name.casefold(), {
            "name": unit_name,
            "symbol": UNIT_SYMBOLS.get(unit_name.casefold(), unit_name[:5].lower()),
        })
        source_key = f"ingredient:{source_id}"
        if code and code not in ingredient_code_to_key:
            ingredient_code_to_key[code] = source_key
        ingredients.append({
            "source_key": source_key,
            "source_code": code,
            "name": display_name(scalar(fields.get("nombre"))),
            "unit_name": unit_name,
            "is_active": str(clean(scalar(fields.get("status", "1"))) == "1").lower(),
            "barcode": clean(scalar(fields.get("codigoBarra"))),
            "quantity": f"{number(scalar(fields.get('cantidad'))):.3f}",
            "cost": f"{number(scalar(fields.get('costo'))):.4f}",
            "source_payload": compact(record),
        })

    _, recipe_source = workbook_rows(official / "recipe_ingredients.xlsx")
    recipes: list[dict[str, Any]] = []
    for row in recipe_source:
        item_code = clean(row["Item Code (*)"])
        ingredient_code = clean(row["Product Code (*)"])
        quantity = number(row["Quantity (*)"])
        unit_name = clean(row["Unit Name (referential only)"]) or "Unidad"
        if item_code in product_code_to_key and ingredient_code in ingredient_code_to_key and quantity > 0:
            units_by_name.setdefault(unit_name.casefold(), {"name": unit_name, "symbol": UNIT_SYMBOLS.get(unit_name.casefold(), unit_name[:5].lower())})
            recipes.append({
                "product_source_key": product_code_to_key[item_code],
                "ingredient_source_key": ingredient_code_to_key[ingredient_code],
                "quantity": f"{quantity:.3f}",
                "unit_name": unit_name,
            })

    modifiers: list[dict[str, Any]] = []
    product_modifiers: list[dict[str, Any]] = []
    for row in modifier_source:
        code = clean(row["Code (If it is new, you can leave this field empty, fill if you want to modify)"])
        source_key = f"modifier:{code}"
        modifiers.append({
            "source_key": source_key,
            "code": code,
            "name": display_name(row["Modifier Name"]),
            "modifier_type": integer(row["Modifier Type (1: Optional, 2: Forced, 3: Multiple)"]) or 1,
            "display_order": integer(row["Order"]) or 0,
            "min_selections": integer(row["Min"]) or 0,
            "max_selections": integer(row["Max"]) or 0,
            "allow_repeat": str(yes(row["Allows Repeat? (Yes/No)"])).lower(),
            "is_active": str(yes(row["Active? (Yes/No)"])).lower(),
            "source_payload": compact(row),
        })
        assigned = clean(row["Item code to which you want to Assign (in case of assigning to several items, you can add them separating them by comma (,). EX: A1,A2,A3)"])
        for item_code in re.split(r"\s*,\s*", assigned):
            if item_code in product_code_to_key:
                product_modifiers.append({"product_source_key": product_code_to_key[item_code], "modifier_source_key": source_key})

    modifier_options: list[dict[str, Any]] = []
    option_code_to_key: dict[str, str] = {}
    for row in option_source:
        modifier_code = clean(row["Modifier code"])
        option_code = clean(row["Modifier option code"])
        source_key = f"modifier-option:{option_code}"
        option_code_to_key[option_code] = source_key
        modifier_options.append({
            "source_key": source_key,
            "modifier_source_key": f"modifier:{modifier_code}",
            "code": option_code,
            "name": display_name(row["Modifier option name"]),
            "sale_price": f"{number(row['Sale price']):.2f}",
            "display_order": integer(row["Order (optional)"]) or 0,
            "is_active": str(yes(row["Active? (Yes/No)"])).lower(),
            "source_payload": compact(row),
        })

    _, modifier_recipe_source = workbook_rows(official / "modifier_recipes.xlsx")
    modifier_ingredients: list[dict[str, Any]] = []
    for row in modifier_recipe_source:
        option_code = clean(row["Modifier option code"])
        ingredient_code = clean(row["Product code"])
        quantity = number(row["Recipe quantity"])
        unit_name = clean(row["Recipe unit name"]) or "Unidad"
        if option_code in option_code_to_key and ingredient_code in ingredient_code_to_key and quantity > 0:
            units_by_name.setdefault(unit_name.casefold(), {"name": unit_name, "symbol": UNIT_SYMBOLS.get(unit_name.casefold(), unit_name[:5].lower())})
            modifier_ingredients.append({
                "option_source_key": option_code_to_key[option_code],
                "ingredient_source_key": ingredient_code_to_key[ingredient_code],
                "quantity": f"{quantity:.3f}",
                "unit_name": unit_name,
                "order_type_code": clean(row["Order type code"]),
            })

    _, customer_source = workbook_rows(official / "customers.xlsx")
    customers: list[dict[str, Any]] = []
    for index, row in enumerate(customer_source, start=1):
        identification = clean(row["No. Identificación"])
        first = clean(row["Nombres"])
        last = clean(row["Apellidos"])
        phones = [clean(row.get(f"Teléfono {i}")) for i in (1, 2, 3)]
        phones = [phone for phone in phones if phone]
        source_key = identification or f"row:{index}"
        customers.append({
            "source_key": source_key,
            "identification": identification,
            "check_digit": clean(row["Digito Verificador"]),
            "first_name": first,
            "last_name": last,
            "full_name": clean(f"{first} {last}") or f"Cliente Invu {index}",
            "phone": phones[0] if phones else "",
            "phones": compact(phones),
            "email": clean(row["Correo electrónico"]),
            "account_status": clean(row["Estado de Cuenta"]),
            "credit_limit": f"{number(row['Límite de Crédito']):.2f}",
            "address": clean(row["Dirección"]),
            "is_active": str(yes(row["Active?"])).lower(),
            "source_payload": compact(row),
        })

    _, supplier_source = workbook_rows(official / "suppliers.xlsx")
    suppliers: list[dict[str, Any]] = []
    for index, row in enumerate(supplier_source, start=1):
        code = clean(row["Code"])
        phones = [clean(row.get(f"Phone number {i}")) for i in (1, 2, 3)]
        suppliers.append({
            "source_key": code or f"row:{index}",
            "source_code": code,
            "name": clean(row["Supplier Name"]),
            "contact": clean(row["Contact name"]),
            "phone": next((phone for phone in phones if phone), ""),
            "email": clean(row["Email"]),
            "notes": clean(" | ".join(filter(None, [clean(row["Ruc"]), clean(row["Payment Address"]), clean(row["Shipping Address"])]))),
            "is_active": str(yes(row["Active? (Yes/No)"])).lower(),
            "source_payload": compact(row),
        })

    _, employee_source = workbook_rows(official / "employees.xlsx")
    employees: list[dict[str, Any]] = []
    for index, source_row in enumerate(employee_source, start=1):
        row = dict(source_row)
        row.pop("Pin", None)  # Never persist plaintext source PINs.
        code = clean(row["ID Code (Only fill if you need to update)"])
        employees.append({
            "source_key": code or f"row:{index}",
            "source_code": code,
            "full_name": clean(f"{clean(row['Name'])} {clean(row['Lastname'])}"),
            "position": clean(row["Employee Rol"]),
            "hourly_rate": "0.00",
            "is_active": str(yes(row["Active? (Yes/No)"])).lower(),
            "source_payload": compact(row),
        })

    sales_source = json.loads((source / "structured" / "closed_orders.json").read_text(encoding="utf-8"))
    sales: list[dict[str, Any]] = []
    for row in sales_source:
        source_id = integer(row.get("id"))
        sales.append({
            "source_key": f"sale:{source_id}",
            "source_id": source_id or "",
            "order_label": plain_html(row.get("orders")),
            "invoice": clean(row.get("invoice")),
            "customer_text": plain_html(row.get("customer")),
            "opened_by": plain_html(row.get("opened_by")),
            "closed_by": plain_html(row.get("closed_by")),
            "table_text": plain_html(row.get("table")),
            "opened_at": iso(row.get("closed_from")),
            "closed_at": iso(row.get("closed_to")),
            "order_type": plain_html(row.get("order_type")),
            "subtotal": f"{number(row.get('subtotal')):.2f}",
            "item_discount": f"{number(row.get('item_discount')):.2f}",
            "order_discount": f"{number(row.get('order_discount')):.2f}",
            "tax": f"{number(row.get('tax')):.2f}",
            "total_tips": f"{number(row.get('total_tips')):.2f}",
            "total": f"{number(row.get('total')):.2f}",
            "cash_paid": f"{number(row.get('cash_paid')):.2f}",
            "card_paid": f"{number(row.get('card_paid')):.2f}",
            "cheque_paid": f"{number(row.get('cheque_paid')):.2f}",
            "other_paid": f"{number(row.get('other_paid')):.2f}",
            "source_payload": compact(row),
        })

    deleted_source = json.loads((source / "structured" / "deleted_orders.json").read_text(encoding="utf-8"))
    deleted: list[dict[str, Any]] = []
    for row in deleted_source:
        source_id = integer(row.get("id"))
        deleted.append({
            "source_key": f"deleted:{source_id}",
            "source_id": source_id or "",
            "order_label": plain_html(row.get("orders")),
            "closed_by": plain_html(row.get("closed_by")),
            "created_at_invu": iso(row.get("create_date")),
            "description": plain_html(row.get("description")),
            "deleted_at": iso(row.get("delete_date")),
            "deleted_by": plain_html(row.get("deleted_by")),
            "source_payload": compact(row),
        })

    _, purchase_source = workbook_rows(official / "purchase_orders.xlsx")
    purchases: list[dict[str, Any]] = []
    for index, row in enumerate(purchase_source, start=1):
        code = clean(row["PO Code"])
        purchases.append({
            "source_key": code or f"row:{index}",
            "po_code": code,
            "supplier_text": clean(row["Supplier"]),
            "creation_date": iso(row["Creation Date"]),
            "po_date": iso(row["PO Date"]),
            "tax": f"{number(row['Tax']):.2f}",
            "discount": f"{number(row['Discount']):.2f}",
            "import_cost": f"{number(row['Import Cost']):.2f}",
            "total": f"{number(row['Total']):.2f}",
            "invoice_number": clean(row["Invoice Number"]),
            "created_by_text": clean(row["Created By"]),
            "status": clean(row["Status"]),
            "notes": clean(row["Notes"]),
            "source_payload": compact(row),
        })

    datasets = {
        "products": products,
        "units": sorted(units_by_name.values(), key=lambda row: row["name"].casefold()),
        "ingredients": ingredients,
        "recipe_components": recipes,
        "modifiers": modifiers,
        "modifier_options": modifier_options,
        "product_modifiers": product_modifiers,
        "modifier_option_ingredients": modifier_ingredients,
        "customers": customers,
        "suppliers": suppliers,
        "employees": employees,
        "legacy_sales": sales,
        "legacy_deleted_orders": deleted,
        "legacy_purchase_orders": purchases,
    }

    for name, rows in datasets.items():
        write_csv(output / f"{name}.csv", rows)

    product_names = [row["name"].casefold() for row in products]
    if len(product_names) != len(set(product_names)):
        raise RuntimeError("Generated product names are not unique")

    manifest: dict[str, Any] = {
        "source": "Invu POS",
        "contains_client_data": True,
        "counts": {name: len(rows) for name, rows in datasets.items()},
        "rules": {
            "zero_price_parent_items": "kept inactive; positive modifier options become active variants",
            "employee_pin": "discarded and never written to bundle",
            "historical_orders": "stored in legacy tables until line details can be recovered",
        },
        "files": {},
    }
    for path in sorted(output.glob("*.csv")):
        raw = path.read_bytes()
        manifest["files"][path.name] = {"bytes": len(raw), "sha256": hashlib.sha256(raw).hexdigest()}
    manifest_raw = compact(manifest).encode("utf-8")
    manifest["bundle_sha256"] = hashlib.sha256(manifest_raw).hexdigest()
    (output / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path, help="FullChina-Invu-Migration-Export directory")
    parser.add_argument("output", type=Path, help="Private output directory")
    args = parser.parse_args()
    print(json.dumps(build(args.source.resolve(), args.output.resolve()), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
